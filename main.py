import traceback
from openpyxl import load_workbook


def main():
    workbook = load_workbook('data.xlsx', read_only=True, data_only=True)

    worksheet = workbook.active
    print(worksheet.max_row)
    print(worksheet.max_column)
    print(worksheet['D1'].value)
    # for column_number in range(1, worksheet.max_column+1):
    #     print(worksheet.cell(row=1, column=column_number), worksheet.cell(row=1, column=column_number).value)

    # for row in worksheet.rows:
    #     for cell in row:
    #         print(cell.value)
    # print(*tuple(worksheet.rows.value), sep='\n')
    table = [[value for value in row] for row in worksheet.values]
    print(*table, sep='\n')
    # for row in worksheet.values:
    #     print()
    #     for value in row:
    #         print(value)

if __name__ == '__main__':
    try:
        main()
    except Exception as error:
        print('Возникла ошибка при выполнении:', error)
        traceback.print_exception(error)
    else:
        print('Выполнение закончено без ошибок.')
